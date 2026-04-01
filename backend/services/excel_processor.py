"""
Excel Processor Service - Batch process questionnaires through the RAG engine.
"""
import uuid
from pathlib import Path
from typing import Dict, Optional
import openpyxl
from services import rag_engine

EXPORTS_DIR = Path(__file__).parent.parent / "data" / "exports"


def process_excel_questionnaire(
    file_path: str,
    question_column: str,
    answer_column: str,
    start_row: int = 2,
    confidence_column: Optional[str] = None,
    source_column: Optional[str] = None,
) -> Dict:
    """Read questions from Excel, query RAG engine, write answers back."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    results, total, answered = [], 0, 0

    q_col = openpyxl.utils.column_index_from_string(question_column)
    a_col = openpyxl.utils.column_index_from_string(answer_column)
    conf_col = openpyxl.utils.column_index_from_string(confidence_column) if confidence_column else None
    src_col = openpyxl.utils.column_index_from_string(source_column) if source_column else None

    for row_idx in range(start_row, ws.max_row + 1):
        question = str(ws.cell(row=row_idx, column=q_col).value or "").strip()
        if not question:
            continue
        total += 1

        try:
            result = rag_engine.query_knowledge_base(question)
            ws.cell(row=row_idx, column=a_col, value=result["answer"])
            if conf_col:
                ws.cell(row=row_idx, column=conf_col, value=result["confidence"].upper())
            if src_col:
                ws.cell(row=row_idx, column=src_col, value=", ".join(s["filename"] for s in result["sources"]))
            answered += 1
            results.append({"row": row_idx, "question": question[:100], "answer": result["answer"][:200], "confidence": result["confidence"], "status": "success"})
        except Exception as e:
            ws.cell(row=row_idx, column=a_col, value=f"[ERROR] {str(e)}")
            results.append({"row": row_idx, "question": question[:100], "answer": f"Error: {str(e)}", "confidence": "error", "status": "error"})

    output_filename = f"answered_{uuid.uuid4().hex[:8]}_{Path(file_path).name}"
    wb.save(str(EXPORTS_DIR / output_filename))
    wb.close()
    return {"output_path": str(EXPORTS_DIR / output_filename), "output_filename": output_filename, "total_questions": total, "answered": answered, "results": results}
